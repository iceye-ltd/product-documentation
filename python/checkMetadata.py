#!/usr/bin/env python3
# (c) ICEYE Oy 2020,2021,2022
# Permission is hereby granted, free of charge, to any
# person obtaining a copy of this software and associated
# documentation files (the "Software") to deal in the
# Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish
# distribute, sublicense,
# and/or sell copies of the Software, and to permit 
# persons to whom the Software is furnished to do so,
# subject to the following conditions:
#
# The above copyright notice and this permission notice 
# shall be included in all copies or substantial portions
# of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF
# ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
# TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A
# PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF
# CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN
# CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SORTWARE.
from copy import deepcopy
from typing import ValuesView
import pandas as pd
import xmltodict
import os
from osgeo import gdal
import h5py
import argparse
import sys
from openpyxl import load_workbook

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

metadataMaster = 'metadataMaster/ICEYE_Product_Metadata.xlsx'

def IsSpecSupported(specVersion):
    global metadataMaster

    sheetName = 'Version '+specVersion
    wb = load_workbook(metadataMaster, read_only=True)   # open an Excel file and return a workbook
    if sheetName not in wb.sheetnames:
        return False
    
    return True

def checkAmplitudeXML(fileName,verbose=False):
    global metadataMaster
    
    if not (os.path.isfile(metadataMaster) and os.access(metadataMaster, os.R_OK)):
        print(f'File {metadataMaster} is not readable or does not exist')
        return False, [], {}

    if not (os.path.isfile(fileName) and os.access(fileName, os.R_OK)):
        print(f'File {fileName} is not readable or does not exist')
        return False, [], {}

    ampxml = open(fileName, 'r').read()  # Read data
    xmlDict = xmltodict.parse(ampxml)    # Parse XML
    xmlDict = dict((k.lower(), v) for k,v in xmlDict['Metadata'].items())
    residuals = deepcopy(xmlDict)         # dict that we remove found items from to find whats left

    specVersion = xmlDict['spec_version']
    sheetName = 'Version '+specVersion

    if not IsSpecSupported(specVersion):
        print(f'Amplitude xml file is version {bcolors.FAIL}{specVersion}{bcolors.ENDC} which does not exist in metatdata master file')
        return False, [], {}

    df = pd.read_excel(metadataMaster, sheet_name=sheetName)

    if verbose :
        print('Amplitude image specification version is',specVersion)
        print('Amplitude image processed with processor version',xmlDict['processor_version'])

    # only scan mode uses beam parameters so remove if not scan mode
    am = xmlDict['acquisition_mode']
    if verbose :
        print('Acquisition mode',am)

    if am != 'scan' :
        ind = df[df['Format Metadata Elements'] == 'scan_beams'].index[0]
        df.drop(labels=[ind], axis=0, inplace=True)

    missingEntries = []
    for tag in df['Format Metadata Elements'] :
        dowefind = False
        if tag.lower() != 'end' and tag.lower() != 'nan':
            line = df[df['Format Metadata Elements'] == tag].index[0]
            if df['AMP-XML'][line] == 'y' or  df['AMP-XML'][line] == 'Y':
                dowefind = True

            if dowefind :
                if tag.lower() not in xmlDict :
                    missingEntries.append(tag)
                else:
                    residuals.pop(tag.lower())

    if len(missingEntries) != 0 :
        if verbose :
            print('The following metadata elements are missing from',fileName)
            for e in missingEntries:
                print(e)
        return False, missingEntries, residuals
    else :
        return True, [], residuals

def checkSLCXML(fileName,verbose=False):
    global metadataMaster

    if not (os.path.isfile(metadataMaster) and os.access(metadataMaster, os.R_OK)):
        print(f'File {metadataMaster} is not readable or does not exist')
        return False, [], {}

    if not (os.path.isfile(fileName) and os.access(fileName, os.R_OK)):
        print(f'File {fileName} is not readable or does not exist')
        return False, [], {}

    slcxml = open(fileName, 'r').read()  # Read data
    xmlDict = xmltodict.parse(slcxml)  # Parse XML
    xmlDict = dict((k.lower(), v) for k,v in xmlDict['Metadata'].items())
    residuals = deepcopy(xmlDict)         # dict that we remove found items from to find whats left

    specVersion = xmlDict['spec_version']
    sheetName = 'Version '+specVersion
    if not IsSpecSupported(specVersion):
        print(f'SLC xml file is version {bcolors.FAIL}{specVersion}{bcolors.ENDC} which does not exist in metatdata master file')
        return False, [], {}

    df = pd.read_excel(metadataMaster, sheet_name=sheetName)

    if verbose :
        print('SLC image specification version is',specVersion)
        print('SLC image processed with processor version',xmlDict['processor_version'])

    # only scan mode uses beam parameters so remove if not scan mode
    am = xmlDict['acquisition_mode']
    if verbose :
        print('Acquisition mode',am)

    if am != 'scan' :
        ind = df[df['Format Metadata Elements'] == 'scan_beams'].index[0]
        df.drop(labels=[ind], axis=0, inplace=True)

    missingEntries = []
    for tag in df['Format Metadata Elements'] :
        dowefind = False
        if tag.lower() != 'end' and tag.lower() != 'nan':
            line = df[df['Format Metadata Elements'] == tag].index[0]
            if df['SLC-XML'][line] == 'y' or  df['SLC-XML'][line] == 'Y':
                dowefind = True

            if dowefind :
                if tag.lower() not in xmlDict :
                    missingEntries.append(tag)
                else:
                    residuals.pop(tag.lower())

    if len(missingEntries) != 0 :
        if verbose :
            print('The following metadata elements are missing from',fileName)
            for e in missingEntries:
                print(e)
        return False, missingEntries, residuals
    else :
        return True, [], residuals

def checkAmpGeotif(fileName,verbose=False):
    global metadataMaster

    if not (os.path.isfile(metadataMaster) and os.access(metadataMaster, os.R_OK)):
        print(f'File {metadataMaster} is not readable or does not exist')
        return False, [], {}

    if not (os.path.isfile(fileName) and os.access(fileName, os.R_OK)):
        print(f'File {fileName} is not readable or does not exist')
        return False, [], {}

    f=gdal.Open(fileName)
    gtifdict=f.GetMetadata()
    residuals = deepcopy(gtifdict)         # dict that we remove found items from to find whats left

    specVersion = gtifdict['SPEC_VERSION']
    sheetName = 'Version '+specVersion
    if not IsSpecSupported(specVersion):
        print(f'Amplitude geotif file is version {bcolors.FAIL}{specVersion}{bcolors.ENDC} which does not exist in metatdata master file')
        return False, [], {}

    df = pd.read_excel(metadataMaster, sheet_name=sheetName)

    if verbose :
        print('amplitude geotif image specification version is',specVersion)
        print('amplitude geotif image processed with processor version',gtifdict['PROCESSOR_VERSION'])


    # geotif files contain RPC data as a separate metadata block
    rpc=f.GetMetadata('RPC')
    if rpc :
        ind = df[df['Format Metadata Elements'] == 'RPC'].index[0]
        df.drop(labels=[ind], axis=0, inplace=True)

    # only scan mode uses beam parameters so remove if not scan mode
    am = gtifdict['ACQUISITION_MODE']
    if verbose :
        print('Acquisition mode',am)

    if am != 'scan' :
        ind = df[df['Format Metadata Elements'] == 'scan_beams'].index[0]
        df.drop(labels=[ind], axis=0, inplace=True)

    missingEntries = []
    for tag in df['Format Metadata Elements'] :
        dowefind = False
        if tag.lower() != 'end' and tag.lower() != 'nan':
            line = df[df['Format Metadata Elements'] == tag].index[0]
            if df['GEOTIF'][line] == 'y' or  df['GEOTIF'][line] == 'Y':
                dowefind = True

            if dowefind :
                if tag.upper()  not in gtifdict:
                    missingEntries.append(tag)
                else:
                    residuals.pop(tag.upper())

    if len(missingEntries) != 0 :
        if verbose :
            print('The following metadNata elements are missing from',fileName)
            for e in missingEntries:
                print(e)
        return False, missingEntries, residuals
    else :
        return True, [], residuals


def checkSlcHdf(fileName,verbose=False):
    global metadataMaster

    if not (os.path.isfile(metadataMaster) and os.access(metadataMaster, os.R_OK)):
        print(f'File {metadataMaster} is not readable or does not exist')
        return False, [], {}

    if not (os.path.isfile(fileName) and os.access(fileName, os.R_OK)):
        print(f'File {fileName} is not readable or does not exist')
        return False, [], {}

    f=h5py.File(fileName)
    slcdict=f.keys()
    residuals = []
    for k in slcdict :
        residuals.append(k)

    specVersion = str(f['spec_version'][()])
    sheetName = 'Version '+specVersion
    if not IsSpecSupported(specVersion):
        print(f'SLC hdf file is version {bcolors.FAIL}{specVersion}{bcolors.ENDC} which does not exist in metatdata master file')
        return False, [], {}

    df = pd.read_excel(metadataMaster, sheet_name=sheetName)

    if verbose :
        print('SLC HDF5 image specification version is',specVersion)
        print('SLC HDF5 image processed with processor version',f['processor_version'][()])


    # only scan mode uses beam parameters so remove if not scan mode
    am = (f['acquisition_mode'][()]).decode('ascii')
    if verbose :
        print('Acquisition mode',am)

    if am != 'scan' :
        ind = df[df['Format Metadata Elements'] == 'scan_beams'].index[0]
        df.drop(labels=[ind], axis=0, inplace=True)

    missingEntries = []
    for tag in df['Format Metadata Elements'] :
        dowefind = False
        if str(tag).lower() != 'end' and str(tag).lower() != 'nan':
            line = df[df['Format Metadata Elements'] == tag].index[0]
            if df['HDF5'][line] == 'y' or  df['HDF5'][line] == 'Y':
                dowefind = True

            if dowefind :
                if tag not in slcdict:
                    missingEntries.append(tag)
                else:
                    residuals.remove(tag)

    if len(missingEntries) != 0 :
        if verbose :
            print('The following metadNata elements are missing from',fileName)
            for e in missingEntries:
                print(e)
        return False, missingEntries, residuals
    else :
        return True, [], residuals


def dumpMDTables(version=2.3):
    global metadataMaster

    if not (os.path.isfile(metadataMaster) and os.access(metadataMaster, os.R_OK)):
        print(f'File {metadataMaster} is not readable or does not exist')
        return False
    
    specVersion = str(version)
    sheetName = 'Version '+specVersion
    df = pd.read_excel(metadataMaster, sheet_name=sheetName)
    for c in df.columns :
        print(f'| {c}',end='')
    print('|')
    for c in df.columns :
        print(f'|-----',end='')
    print('|')

    for tag in df['Format Metadata Elements'] :
        row = df[df['Format Metadata Elements'] == tag].index[0]
        for c in df.columns :
            txt = str(df[c][row]).replace('\n', ' ')
            txt = txt.replace('nan',' ')
            if c == 'HDF5' or c == 'SLC-XML' or c=='GEOTIF' or c == 'AMP-XML' :
                txt=txt.lower()
                if txt.isspace() :
                    txt = '<span style=\"color:red\">:material-close:'
                elif txt == 'n':
                    txt = txt.replace('n','<span style=\"color:red\">:material-close:')
                elif txt == 'y':
                    txt = txt.replace('y','<span style=\"color:green\">:material-check:')

            if c == 'Format Metadata Elements' :
                print(f'|`{txt}`',end='')
            else:
                print(f'|{txt}',end='')
        print('|')

    return True


if __name__ == "__main__":
    ''' 
    Program to compare the metadata tags in an ICEYE image to a master list of metadata
    '''
    verb = False

    parser = argparse.ArgumentParser(description='Check ICEYE Product Metadata')
    parser.add_argument('-f','--files',type=str,nargs='+')
    parser.add_argument('-v','--verbose', help='print useful output wile running',action='store_true')
    parser.add_argument('-d','--dump',help='dump metadata as a markdown table for version number',type=float)
    parser.add_argument('-m','--master',help='set path to metadata master document')
    parser.add_argument('-i','--ignore',help='ignore files than do not have known ICEYE metadata',action='store_true')
    parser.add_argument('-x','--extra',help='print out any extra metadata elements not found in the specification',action='store_true')

    args=parser.parse_args()

    if args.verbose :
        verb = True
 
    if args.master :
        metadataMaster = args.master
        if not (os.path.isfile(metadataMaster) and os.access(metadataMaster, os.R_OK)):
            sys.exit(f'File {metadataMaster} is not readable or does not exist')

    if args.dump :
        if args.dump != 2.4 and args.dump != 2.3 :
            print(f'Available metadata versions are 2.3 and 2.4')
            exit(1)
        dumpMDTables(version=args.dump)

    else:
        for file in args.files :
            fname, extension = os.path.splitext(file)
            if extension == '.xml' :
                prodtype = (fname.split("_")[2]).upper()
                if prodtype == 'GRD' :
                    ampxmlfile = file
                    status, missing, residuals = checkAmplitudeXML(ampxmlfile,verbose=verb)
                    if not status :
                        print(f'{ampxmlfile}  \t: {bcolors.FAIL}Failed{bcolors.ENDC}')
                    else :
                        print(f'{ampxmlfile}  \t: {bcolors.OKGREEN}OK{bcolors.ENDC}')
                    if args.extra and len(residuals) > 0 :
                        print(f'Extra metadata elements in {ampxmlfile}')
                        for k in residuals : 
                            print(f'    {k}')

                elif prodtype == 'SLC' :
                    slcxmlfile = file
                    status, missin, residuals = checkSLCXML(slcxmlfile,verbose=verb)
                    if not status :
                        print(f'{slcxmlfile}  \t: {bcolors.FAIL}Failed{bcolors.ENDC}')
                    else :
                        print(f'{slcxmlfile}  \t: {bcolors.OKGREEN}OK{bcolors.ENDC}')
                    if args.extra and len(residuals) > 0:
                        print(f'Extra metadata elements in {slcxmlfile}')
                        for k in residuals : 
                            print(f'    {k}')

                else:
                    if verb :
                        print(f'No known metadata for file : {bcolors.WARNING}{file}{bcolors.ENDC}. Skipping...')
            
            elif extension == '.tif' :
                ampgeotiffile = file 
                status, missing, residuals = checkAmpGeotif(ampgeotiffile,verbose=verb)
                if not status :
                    print(f'{ampgeotiffile}  \t: {bcolors.FAIL}Failed{bcolors.ENDC}')
                else :
                    print(f'{ampgeotiffile}  \t: {bcolors.OKGREEN}OK{bcolors.ENDC}')
                if args.extra and len(residuals) > 0:
                    print(f'Extra metadata elements in {ampgeotiffile}')
                    for k in residuals : 
                        print(f'    {k}')

            elif extension == '.h5' :
                slchdffile = file
                status, missing, residuals = checkSlcHdf(slchdffile,verbose=verb)
                if not status :
                    print(f'{slchdffile} \t: {bcolors.FAIL}Failed{bcolors.ENDC}')
                else :
                    print(f'{slchdffile} \t: {bcolors.OKGREEN}OK{bcolors.ENDC}')
                if args.extra and len(residuals) > 0:
                    print(f'Extra metadata elements in {slchdffile}')
                    for k in residuals : 
                        print(f'    {k}')

            else :
                if not args.ignore :
                    print(f'No metadata for file extension {bcolors.WARNING}{file}{bcolors.ENDC} Skipping...')